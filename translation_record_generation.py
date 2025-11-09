import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
from datetime import datetime


class TranslationRecordGenerator:
    """
    Generates an Excel record comparing original and translated PowerPoint content.
    
    Creates a detailed spreadsheet showing:
    - Original English text
    - Translated text
    - Metadata (slide number, element type, location, etc.)
    """
    
    def __init__(self, extraction_json_path: str, translation_json_path: str):
        """
        Initialize the record generator.
        
        Args:
            extraction_json_path: Path to extraction JSON (original content)
            translation_json_path: Path to translation JSON (translated content)
        """
        self.extraction_path = extraction_json_path
        self.translation_path = translation_json_path
        
        # Load both JSONs
        print(f"Loading extraction JSON: {extraction_json_path}")
        with open(extraction_json_path, 'r', encoding='utf-8') as f:
            self.extraction_data = json.load(f)
        
        print(f"Loading translation JSON: {translation_json_path}")
        with open(translation_json_path, 'r', encoding='utf-8') as f:
            self.translation_data = json.load(f)
        
        # Get target language
        self.target_language = self.translation_data.get('target_language', 'Unknown')
        self.is_rtl = self.translation_data.get('is_rtl', False)
        
        # Statistics
        self.stats = {
            "total_records": 0,
            "text_runs": 0,
            "tables": 0,
            "charts": 0,
            "speaker_notes": 0,
            "smartart": 0
        }
        
        print(f"✓ Loaded extraction: {self.extraction_data['total_slides']} slides")
        print(f"✓ Loaded translation: {self.translation_data['total_slides']} slides")
        print(f"✓ Target language: {self.target_language}")
        if self.is_rtl:
            print(f"✓ RTL mode: ENABLED")
    
    def sanitize_text(self, text):
        """
        Sanitize text to remove illegal characters for Excel.
        
        Excel does not allow certain control characters (0x00-0x1F except tab, newline, carriage return).
        
        Args:
            text: Text string to sanitize
            
        Returns:
            Sanitized text string safe for Excel
        """
        if not text:
            return text
        
        if not isinstance(text, str):
            return text
        
        # Remove illegal XML characters for Excel
        # Keep: tab (0x09), newline (0x0A), carriage return (0x0D)
        # Remove: all other control characters (0x00-0x1F) and 0x7F-0x9F
        illegal_chars = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]')
        sanitized = illegal_chars.sub('', text)
        
        # Also remove any other problematic characters
        # Remove vertical tab, form feed, etc.
        sanitized = sanitized.replace('\v', '').replace('\f', '')
        
        return sanitized
    
    def create_workbook(self):
        """Create and style the Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Translation Record"
        
        # Define headers
        headers = [
            "Record ID",
            "Slide Number",
            "Element Type",
            "Element Name",
            "Location (Top, Left)",
            "Original Text (English)",
            "Translated Text",
            "Char Count (Original)",
            "Char Count (Translated)",
            "Length Change %",
            "Font Name",
            "Font Size",
            "Bold",
            "Italic",
            "Underline",
            "Font Color",
            "Text Alignment",
            "Is Bulleted",
            "Bullet Type",
            "Placeholder Type",
            "Shape Width",
            "Shape Height",
            "Background Color",
            "Has Shadow",
            "Text Direction",
            "Notes"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = {
            'A': 10,  # Record ID
            'B': 12,  # Slide Number
            'C': 15,  # Element Type
            'D': 20,  # Element Name
            'E': 25,  # Location
            'F': 50,  # Original Text
            'G': 50,  # Translated Text
            'H': 12,  # Char Count Original
            'I': 12,  # Char Count Translated
            'J': 12,  # Length Change %
            'K': 15,  # Font Name
            'L': 10,  # Font Size
            'M': 8,   # Bold
            'N': 8,   # Italic
            'O': 10,  # Underline
            'P': 15,  # Font Color
            'Q': 15,  # Text Alignment
            'R': 12,  # Is Bulleted
            'S': 15,  # Bullet Type
            'T': 18,  # Placeholder Type
            'U': 12,  # Shape Width
            'V': 12,  # Shape Height
            'W': 18,  # Background Color
            'X': 12,  # Has Shadow
            'Y': 15,  # Text Direction
            'Z': 30   # Notes
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        return wb, ws
    
    def add_record(self, ws, row_num, record_data):
        """
        Add a translation record to the worksheet.
        
        Args:
            ws: Worksheet object
            row_num: Row number to add record
            record_data: Dictionary with record information
        """
        # Cell styling
        text_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Alternate row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Add data with all metadata columns - SANITIZE TEXT VALUES
        columns = [
            ('A', record_data.get('record_id', ''), center_alignment),
            ('B', record_data.get('slide_number', ''), center_alignment),
            ('C', self.sanitize_text(record_data.get('element_type', '')), center_alignment),
            ('D', self.sanitize_text(record_data.get('element_name', '')), text_alignment),
            ('E', self.sanitize_text(record_data.get('location', '')), text_alignment),
            ('F', self.sanitize_text(record_data.get('original_text', '')), text_alignment),
            ('G', self.sanitize_text(record_data.get('translated_text', '')), text_alignment),
            ('H', record_data.get('char_count_original', ''), center_alignment),
            ('I', record_data.get('char_count_translated', ''), center_alignment),
            ('J', self.sanitize_text(record_data.get('length_change_percent', '')), center_alignment),
            ('K', self.sanitize_text(record_data.get('font_name', '')), center_alignment),
            ('L', self.sanitize_text(record_data.get('font_size', '')), center_alignment),
            ('M', self.sanitize_text(record_data.get('bold', '')), center_alignment),
            ('N', self.sanitize_text(record_data.get('italic', '')), center_alignment),
            ('O', self.sanitize_text(record_data.get('underline', '')), center_alignment),
            ('P', self.sanitize_text(record_data.get('font_color', '')), center_alignment),
            ('Q', self.sanitize_text(record_data.get('text_alignment', '')), center_alignment),
            ('R', self.sanitize_text(record_data.get('is_bulleted', '')), center_alignment),
            ('S', self.sanitize_text(record_data.get('bullet_type', '')), center_alignment),
            ('T', self.sanitize_text(record_data.get('placeholder_type', '')), center_alignment),
            ('U', self.sanitize_text(record_data.get('shape_width', '')), center_alignment),
            ('V', self.sanitize_text(record_data.get('shape_height', '')), center_alignment),
            ('W', self.sanitize_text(record_data.get('background_color', '')), center_alignment),
            ('X', self.sanitize_text(record_data.get('has_shadow', '')), center_alignment),
            ('Y', self.sanitize_text(record_data.get('text_direction', '')), center_alignment),
            ('Z', self.sanitize_text(record_data.get('notes', '')), text_alignment)
        ]
        
        for col, value, alignment in columns:
            cell = ws[f"{col}{row_num}"]
            cell.value = value
            cell.alignment = alignment
            cell.fill = fill
            cell.border = border
        
        # Set row height for better readability
        ws.row_dimensions[row_num].height = 30
    
    def extract_text_from_runs(self, runs):
        """Extract concatenated text from runs"""
        if not runs:
            return ""
        return "".join(run.get("text", "") for run in runs)
    
    def extract_metadata_from_element(self, element, run=None, para_format=None):
        """
        Extract comprehensive metadata from an element and its run/paragraph formatting.
        
        Args:
            element: Element dictionary
            run: Specific run dictionary (optional)
            para_format: Paragraph formatting dictionary (optional)
            
        Returns:
            Dictionary with metadata fields
        """
        metadata = {
            "font_name": "",
            "font_size": "",
            "bold": "",
            "italic": "",
            "underline": "",
            "font_color": "",
            "text_alignment": "",
            "is_bulleted": "",
            "bullet_type": "",
            "placeholder_type": "",
            "shape_width": "",
            "shape_height": "",
            "background_color": "",
            "has_shadow": "",
            "text_direction": ""
        }
        
        # Extract from run (font formatting)
        if run:
            metadata["font_name"] = run.get("font_name", "")
            metadata["font_size"] = f"{run.get('font_size', '')}pt" if run.get('font_size') else ""
            metadata["bold"] = "Yes" if run.get("bold") else "No"
            metadata["italic"] = "Yes" if run.get("italic") else "No"
            metadata["underline"] = "Yes" if run.get("underline") else "No"
            
            # Font color
            font_color = run.get("color")
            if font_color:
                if isinstance(font_color, dict):
                    if 'rgb' in font_color:
                        metadata["font_color"] = font_color['rgb']
                    elif 'theme_color' in font_color:
                        metadata["font_color"] = font_color['theme_color']
                else:
                    metadata["font_color"] = str(font_color)
        
        # Extract from paragraph formatting
        if para_format:
            alignment = para_format.get("alignment", "")
            if alignment:
                # Clean up alignment string (e.g., "PP_ALIGN.CENTER" -> "CENTER")
                if "." in str(alignment):
                    metadata["text_alignment"] = str(alignment).split(".")[-1]
                else:
                    metadata["text_alignment"] = str(alignment)
            
            # Bullet information
            bullet_info = para_format.get("bullet_format", {})
            metadata["is_bulleted"] = "Yes" if bullet_info.get("is_bulleted") else "No"
            bullet_type = bullet_info.get("bullet_type", "")
            if bullet_type:
                metadata["bullet_type"] = bullet_type
            
            # Text direction
            text_dir = para_format.get("text_direction", "")
            if text_dir:
                metadata["text_direction"] = text_dir
        
        # Extract from element (shape properties)
        if element:
            # Placeholder info
            placeholder_info = element.get("placeholder_info", {})
            if placeholder_info.get("is_placeholder"):
                ph_type = placeholder_info.get("placeholder_type", "")
                if ph_type and "." in ph_type:
                    metadata["placeholder_type"] = ph_type.split(".")[-1]
                else:
                    metadata["placeholder_type"] = ph_type
            
            # Dimensions
            dimensions = element.get("dimensions", {})
            width = dimensions.get("width", "")
            height = dimensions.get("height", "")
            if width:
                # Convert EMUs to inches (914400 EMUs = 1 inch)
                try:
                    metadata["shape_width"] = f"{width / 914400:.2f} in"
                except:
                    metadata["shape_width"] = str(width)
            if height:
                try:
                    metadata["shape_height"] = f"{height / 914400:.2f} in"
                except:
                    metadata["shape_height"] = str(height)
            
            # Fill/Background color
            fill_info = element.get("fill", {})
            if fill_info:
                solid_color = fill_info.get("solid_color")
                if solid_color:
                    if isinstance(solid_color, dict):
                        if 'rgb' in solid_color:
                            metadata["background_color"] = solid_color['rgb']
                        elif 'theme_color' in solid_color:
                            metadata["background_color"] = solid_color['theme_color']
                    else:
                        metadata["background_color"] = str(solid_color)
            
            # Shadow
            shadow_info = element.get("shadow", {})
            if shadow_info:
                metadata["has_shadow"] = "Yes" if shadow_info.get("has_shadow") else "No"
            
            # Text frame properties (text direction)
            text_frame_props = element.get("text_frame_properties", {})
            if text_frame_props:
                text_dir = text_frame_props.get("text_direction", "")
                if text_dir and not metadata["text_direction"]:
                    metadata["text_direction"] = text_dir
        
        return metadata
    
    def process_text_element(self, original_elem, translated_elem, slide_num, records):
        """
        Process text elements (TextBox, AutoShape) and extract text runs with metadata.
        
        Args:
            original_elem: Original element from extraction JSON
            translated_elem: Translated element from translation JSON
            slide_num: Slide number
            records: List to append records to
        """
        element_type = original_elem.get("element_type", "Unknown")
        element_name = original_elem.get("shape_name", "Unnamed")
        
        # Get position info
        dimensions = original_elem.get("dimensions", {})
        top = dimensions.get('top', 'N/A')
        left = dimensions.get('left', 'N/A')
        # Convert EMUs to inches for readability
        try:
            top_in = f"{top / 914400:.2f} in"
            left_in = f"{left / 914400:.2f} in"
            location = f"Top: {top_in}, Left: {left_in}"
        except:
            location = f"Top: {top}, Left: {left}"
        
        # Check if it's a placeholder
        placeholder_info = original_elem.get("placeholder_info", {})
        if placeholder_info.get("is_placeholder"):
            element_type = f"{element_type} (Placeholder)"
            ph_type = placeholder_info.get("placeholder_type", "")
            if ph_type:
                element_name = f"{element_name} [{ph_type}]"
        
        # Process paragraphs and runs
        original_paragraphs = original_elem.get("paragraphs", [])
        translated_paragraphs = translated_elem.get("paragraphs", [])
        
        for para_idx, (orig_para, trans_para) in enumerate(zip(original_paragraphs, translated_paragraphs)):
            orig_runs = orig_para.get("runs", [])
            trans_runs = trans_para.get("runs", [])
            para_format = orig_para.get("paragraph_formatting", {})
            
            for run_idx, (orig_run, trans_run) in enumerate(zip(orig_runs, trans_runs)):
                original_text = orig_run.get("text", "")
                translated_text = trans_run.get("text", "")
                
                # Skip empty texts
                if not original_text.strip() and not translated_text.strip():
                    continue
                
                # Extract metadata from element, run, and paragraph
                metadata = self.extract_metadata_from_element(original_elem, orig_run, para_format)
                
                # Calculate length change percentage
                orig_len = len(original_text)
                trans_len = len(translated_text)
                if orig_len > 0:
                    length_change = ((trans_len - orig_len) / orig_len) * 100
                    length_change_str = f"{length_change:+.1f}%"
                else:
                    length_change_str = "N/A"
                
                # Build notes
                notes_parts = []
                if metadata["is_bulleted"] == "Yes":
                    notes_parts.append(f"Bullet: {metadata['bullet_type']}")
                if orig_run.get("superscript"):
                    notes_parts.append("Superscript")
                if orig_run.get("subscript"):
                    notes_parts.append("Subscript")
                if orig_run.get("strike"):
                    notes_parts.append("Strikethrough")
                
                notes = ", ".join(notes_parts) if notes_parts else ""
                
                record = {
                    "record_id": self.stats["total_records"] + 1,
                    "slide_number": slide_num,
                    "element_type": element_type,
                    "element_name": element_name,
                    "location": location,
                    "original_text": original_text,
                    "translated_text": translated_text,
                    "char_count_original": orig_len,
                    "char_count_translated": trans_len,
                    "length_change_percent": length_change_str,
                    "font_name": metadata["font_name"],
                    "font_size": metadata["font_size"],
                    "bold": metadata["bold"],
                    "italic": metadata["italic"],
                    "underline": metadata["underline"],
                    "font_color": metadata["font_color"],
                    "text_alignment": metadata["text_alignment"],
                    "is_bulleted": metadata["is_bulleted"],
                    "bullet_type": metadata["bullet_type"],
                    "placeholder_type": metadata["placeholder_type"],
                    "shape_width": metadata["shape_width"],
                    "shape_height": metadata["shape_height"],
                    "background_color": metadata["background_color"],
                    "has_shadow": metadata["has_shadow"],
                    "text_direction": metadata["text_direction"],
                    "notes": notes
                }
                
                records.append(record)
                self.stats["total_records"] += 1
                self.stats["text_runs"] += 1
    
    def process_table(self, original_table, translated_table, slide_num, element_name, element, records):
        """
        Process table elements with metadata.
        
        Args:
            original_table: Original table_data from extraction JSON
            translated_table: Translated table_data from translation JSON
            slide_num: Slide number
            element_name: Name of the table element
            element: The parent element (for metadata extraction)
            records: List to append records to
        """
        original_cells = original_table.get("cells", [])
        translated_cells = translated_table.get("cells", [])
        
        for orig_cell, trans_cell in zip(original_cells, translated_cells):
            row = orig_cell.get("row", "?")
            col = orig_cell.get("column", "?")
            location = f"Cell ({row}, {col})"
            
            # Process each paragraph in cell
            orig_paragraphs = orig_cell.get("paragraphs", [])
            trans_paragraphs = trans_cell.get("paragraphs", [])
            
            for para_idx, (orig_para, trans_para) in enumerate(zip(orig_paragraphs, trans_paragraphs)):
                orig_text = self.extract_text_from_runs(orig_para.get("runs", []))
                trans_text = self.extract_text_from_runs(trans_para.get("runs", []))
                
                # Skip empty cells
                if not orig_text.strip() and not trans_text.strip():
                    continue
                
                # Extract metadata from first run if available
                orig_runs = orig_para.get("runs", [])
                first_run = orig_runs[0] if orig_runs else None
                para_format = orig_para.get("paragraph_formatting", {})
                
                metadata = self.extract_metadata_from_element(element, first_run, para_format)
                
                # Calculate length change
                orig_len = len(orig_text)
                trans_len = len(trans_text)
                if orig_len > 0:
                    length_change = ((trans_len - orig_len) / orig_len) * 100
                    length_change_str = f"{length_change:+.1f}%"
                else:
                    length_change_str = "N/A"
                
                record = {
                    "record_id": self.stats["total_records"] + 1,
                    "slide_number": slide_num,
                    "element_type": "Table Cell",
                    "element_name": element_name,
                    "location": location,
                    "original_text": orig_text,
                    "translated_text": trans_text,
                    "char_count_original": orig_len,
                    "char_count_translated": trans_len,
                    "length_change_percent": length_change_str,
                    "font_name": metadata["font_name"],
                    "font_size": metadata["font_size"],
                    "bold": metadata["bold"],
                    "italic": metadata["italic"],
                    "underline": metadata["underline"],
                    "font_color": metadata["font_color"],
                    "text_alignment": metadata["text_alignment"],
                    "is_bulleted": metadata["is_bulleted"],
                    "bullet_type": metadata["bullet_type"],
                    "placeholder_type": metadata["placeholder_type"],
                    "shape_width": metadata["shape_width"],
                    "shape_height": metadata["shape_height"],
                    "background_color": metadata["background_color"],
                    "has_shadow": metadata["has_shadow"],
                    "text_direction": metadata["text_direction"],
                    "notes": f"Row {row}, Column {col}"
                }
                
                records.append(record)
                self.stats["total_records"] += 1
                self.stats["tables"] += 1
    
    def process_chart(self, original_chart, translated_chart, slide_num, element_name, records):
        """
        Process chart elements with metadata.
        
        Args:
            original_chart: Original chart_data from extraction JSON
            translated_chart: Translated chart_data from translation JSON
            slide_num: Slide number
            element_name: Name of the chart element
            records: List to append records to
        """
        chart_type = original_chart.get('chart_type', 'Unknown')
        
        # Empty metadata for charts (they don't have font formatting)
        empty_metadata = {
            "font_name": "", "font_size": "", "bold": "", "italic": "", 
            "underline": "", "font_color": "", "text_alignment": "", 
            "is_bulleted": "", "bullet_type": "", "placeholder_type": "",
            "shape_width": "", "shape_height": "", "background_color": "",
            "has_shadow": "", "text_direction": ""
        }
        
        # Chart title
        if original_chart.get("title") and translated_chart.get("title"):
            orig_len = len(original_chart["title"])
            trans_len = len(translated_chart["title"])
            length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
            
            record = {
                "record_id": self.stats["total_records"] + 1,
                "slide_number": slide_num,
                "element_type": "Chart Title",
                "element_name": element_name,
                "location": "Chart Title",
                "original_text": original_chart["title"],
                "translated_text": translated_chart["title"],
                "char_count_original": orig_len,
                "char_count_translated": trans_len,
                "length_change_percent": f"{length_change:+.1f}%",
                **empty_metadata,
                "notes": f"Chart Type: {chart_type}"
            }
            records.append(record)
            self.stats["total_records"] += 1
            self.stats["charts"] += 1
        
        # Series names
        orig_series = original_chart.get("series_names", [])
        trans_series = translated_chart.get("series_names", [])
        
        for idx, (orig_name, trans_name) in enumerate(zip(orig_series, trans_series)):
            if orig_name and trans_name:
                orig_len = len(orig_name)
                trans_len = len(trans_name)
                length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
                
                record = {
                    "record_id": self.stats["total_records"] + 1,
                    "slide_number": slide_num,
                    "element_type": "Chart Series",
                    "element_name": element_name,
                    "location": f"Series {idx + 1}",
                    "original_text": orig_name,
                    "translated_text": trans_name,
                    "char_count_original": orig_len,
                    "char_count_translated": trans_len,
                    "length_change_percent": f"{length_change:+.1f}%",
                    **empty_metadata,
                    "notes": f"Chart Type: {chart_type}"
                }
                records.append(record)
                self.stats["total_records"] += 1
                self.stats["charts"] += 1
        
        # Categories (if text)
        orig_categories = original_chart.get("categories", [])
        trans_categories = translated_chart.get("categories", [])
        
        for idx, (orig_cat, trans_cat) in enumerate(zip(orig_categories, trans_categories)):
            if isinstance(orig_cat, str) and isinstance(trans_cat, str):
                orig_len = len(orig_cat)
                trans_len = len(trans_cat)
                length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
                
                record = {
                    "record_id": self.stats["total_records"] + 1,
                    "slide_number": slide_num,
                    "element_type": "Chart Category",
                    "element_name": element_name,
                    "location": f"Category {idx + 1}",
                    "original_text": orig_cat,
                    "translated_text": trans_cat,
                    "char_count_original": orig_len,
                    "char_count_translated": trans_len,
                    "length_change_percent": f"{length_change:+.1f}%",
                    **empty_metadata,
                    "notes": f"Chart Type: {chart_type}"
                }
                records.append(record)
                self.stats["total_records"] += 1
                self.stats["charts"] += 1
    
    def process_smartart(self, original_smartart, translated_smartart, slide_num, records):
        """
        Process SmartArt elements with metadata.
        
        Args:
            original_smartart: Original smartart list from extraction JSON
            translated_smartart: Translated smartart list from translation JSON
            slide_num: Slide number
            records: List to append records to
        """
        empty_metadata = {
            "font_name": "", "font_size": "", "bold": "", "italic": "", 
            "underline": "", "font_color": "", "text_alignment": "", 
            "is_bulleted": "", "bullet_type": "", "placeholder_type": "",
            "shape_width": "", "shape_height": "", "background_color": "",
            "has_shadow": "", "text_direction": ""
        }
        
        for smartart_idx, (orig_smart, trans_smart) in enumerate(zip(original_smartart, translated_smartart)):
            layout_type = orig_smart.get("layout_type", "Unknown")
            
            # Process nodes if available
            orig_nodes = orig_smart.get("nodes", [])
            trans_nodes = trans_smart.get("nodes", [])
            
            if orig_nodes and trans_nodes:
                for node_idx, (orig_node, trans_node) in enumerate(zip(orig_nodes, trans_nodes)):
                    orig_text = orig_node.get("text", "")
                    trans_text = trans_node.get("text", "")
                    
                    if not orig_text.strip() and not trans_text.strip():
                        continue
                    
                    level = orig_node.get("level", "N/A")
                    node_id = orig_node.get("node_id", "")
                    
                    orig_len = len(orig_text)
                    trans_len = len(trans_text)
                    length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
                    
                    record = {
                        "record_id": self.stats["total_records"] + 1,
                        "slide_number": slide_num,
                        "element_type": "SmartArt Node",
                        "element_name": f"SmartArt {smartart_idx + 1}",
                        "location": f"Node {node_idx + 1} (Level {level})",
                        "original_text": orig_text,
                        "translated_text": trans_text,
                        "char_count_original": orig_len,
                        "char_count_translated": trans_len,
                        "length_change_percent": f"{length_change:+.1f}%",
                        **empty_metadata,
                        "notes": f"Layout: {layout_type}, Node ID: {node_id}"
                    }
                    records.append(record)
                    self.stats["total_records"] += 1
                    self.stats["smartart"] += 1
            else:
                # Fallback to texts list
                orig_texts = orig_smart.get("texts", [])
                trans_texts = trans_smart.get("texts", [])
                
                for text_idx, (orig_text, trans_text) in enumerate(zip(orig_texts, trans_texts)):
                    if not orig_text.strip() and not trans_text.strip():
                        continue
                    
                    orig_len = len(orig_text)
                    trans_len = len(trans_text)
                    length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
                    
                    record = {
                        "record_id": self.stats["total_records"] + 1,
                        "slide_number": slide_num,
                        "element_type": "SmartArt Text",
                        "element_name": f"SmartArt {smartart_idx + 1}",
                        "location": f"Text {text_idx + 1}",
                        "original_text": orig_text,
                        "translated_text": trans_text,
                        "char_count_original": orig_len,
                        "char_count_translated": trans_len,
                        "length_change_percent": f"{length_change:+.1f}%",
                        **empty_metadata,
                        "notes": f"Layout: {layout_type}"
                    }
                    records.append(record)
                    self.stats["total_records"] += 1
                    self.stats["smartart"] += 1
    
    def process_speaker_notes(self, original_notes, translated_notes, slide_num, records):
        """
        Process speaker notes with metadata.
        
        Args:
            original_notes: Original speaker_notes from extraction JSON
            translated_notes: Translated speaker_notes from translation JSON
            slide_num: Slide number
            records: List to append records to
        """
        if not original_notes or not translated_notes:
            return
        
        orig_text = original_notes.get("text", "")
        trans_text = translated_notes.get("text", "")
        
        if not orig_text.strip() and not trans_text.strip():
            return
        
        empty_metadata = {
            "font_name": "", "font_size": "", "bold": "", "italic": "", 
            "underline": "", "font_color": "", "text_alignment": "", 
            "is_bulleted": "", "bullet_type": "", "placeholder_type": "",
            "shape_width": "", "shape_height": "", "background_color": "",
            "has_shadow": "", "text_direction": ""
        }
        
        orig_len = len(orig_text)
        trans_len = len(trans_text)
        length_change = ((trans_len - orig_len) / orig_len * 100) if orig_len > 0 else 0
        
        record = {
            "record_id": self.stats["total_records"] + 1,
            "slide_number": slide_num,
            "element_type": "Speaker Notes",
            "element_name": "Notes",
            "location": "Slide Notes",
            "original_text": orig_text,
            "translated_text": trans_text,
            "char_count_original": orig_len,
            "char_count_translated": trans_len,
            "length_change_percent": f"{length_change:+.1f}%",
            **empty_metadata,
            "notes": "Speaker notes for presentation"
        }
        records.append(record)
        self.stats["total_records"] += 1
        self.stats["speaker_notes"] += 1
    
    def generate_records(self):
        """
        Generate all translation records by comparing extraction and translation JSONs.
        
        Returns:
            List of record dictionaries
        """
        print("\nGenerating translation records...")
        print("=" * 80)
        
        records = []
        
        # Process each slide
        original_slides = self.extraction_data.get("slides", [])
        translated_slides = self.translation_data.get("slides", [])
        
        for slide_idx, (orig_slide, trans_slide) in enumerate(zip(original_slides, translated_slides), 1):
            print(f"Processing slide {slide_idx}/{len(original_slides)}...", end=" ")
            
            # Process elements
            orig_elements = orig_slide.get("elements", [])
            trans_elements = trans_slide.get("elements", [])
            
            for orig_elem, trans_elem in zip(orig_elements, trans_elements):
                element_type = orig_elem.get("element_type")
                element_name = orig_elem.get("shape_name", "Unnamed")
                
                if element_type in ["TextBox", "AutoShape"]:
                    self.process_text_element(orig_elem, trans_elem, slide_idx, records)
                
                elif element_type == "Table":
                    orig_table = orig_elem.get("table_data")
                    trans_table = trans_elem.get("table_data")
                    if orig_table and trans_table:
                        self.process_table(orig_table, trans_table, slide_idx, element_name, orig_elem, records)
                
                elif element_type == "Chart":
                    orig_chart = orig_elem.get("chart_data")
                    trans_chart = trans_elem.get("chart_data")
                    if orig_chart and trans_chart:
                        self.process_chart(orig_chart, trans_chart, slide_idx, element_name, records)
            
            # Process SmartArt
            orig_smartart = orig_slide.get("smartart", [])
            trans_smartart = trans_slide.get("smartart", [])
            if orig_smartart and trans_smartart:
                self.process_smartart(orig_smartart, trans_smartart, slide_idx, records)
            
            # Process speaker notes
            orig_notes = orig_slide.get("speaker_notes")
            trans_notes = trans_slide.get("speaker_notes")
            self.process_speaker_notes(orig_notes, trans_notes, slide_idx, records)
            
            print("✓")
        
        print("=" * 80)
        return records
    
    def generate_excel(self, output_path: str):
        """
        Generate the Excel file with translation records.
        
        Args:
            output_path: Path to save the Excel file
        """
        print("\nCreating Excel workbook...")
        
        # Create workbook
        wb, ws = self.create_workbook()
        
        # Generate records
        records = self.generate_records()
        
        print(f"\nWriting {len(records)} records to Excel...")
        
        # Add records to worksheet
        for idx, record in enumerate(records, 2):  # Start from row 2 (after header)
            self.add_record(ws, idx, record)
            
            # Progress indicator
            if idx % 100 == 0:
                print(f"  Written {idx - 1} records...")
        
        # Add summary sheet
        self.add_summary_sheet(wb)
        
        # Save workbook
        print(f"\nSaving Excel file to: {output_path}")
        wb.save(output_path)
        
        # Print statistics
        print("\n" + "=" * 80)
        print("EXCEL GENERATION COMPLETE!")
        print("=" * 80)
        print(f"Output file: {output_path}")
        print(f"Target language: {self.target_language}")
        if self.is_rtl:
            print(f"RTL mode: ENABLED")
        print(f"\nTotal records: {self.stats['total_records']}")
        print(f"  - Text runs: {self.stats['text_runs']}")
        print(f"  - Table cells: {self.stats['tables']}")
        print(f"  - Chart elements: {self.stats['charts']}")
        print(f"  - SmartArt nodes: {self.stats['smartart']}")
        print(f"  - Speaker notes: {self.stats['speaker_notes']}")
        print("=" * 80)
    
    def add_summary_sheet(self, wb):
        """Add a summary sheet with statistics"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Translation Record Summary"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        
        # Metadata
        metadata = [
            ("", ""),
            ("Generation Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Source Files:", ""),
            ("  Extraction JSON:", os.path.basename(self.extraction_path)),
            ("  Translation JSON:", os.path.basename(self.translation_path)),
            ("", ""),
            ("Translation Info:", ""),
            ("  Target Language:", self.target_language),
            ("  RTL Mode:", "Yes" if self.is_rtl else "No"),
            ("  Total Slides:", self.extraction_data['total_slides']),
            ("", ""),
            ("Statistics:", ""),
            ("  Total Records:", self.stats['total_records']),
            ("  Text Runs:", self.stats['text_runs']),
            ("  Table Cells:", self.stats['tables']),
            ("  Chart Elements:", self.stats['charts']),
            ("  SmartArt Nodes:", self.stats['smartart']),
            ("  Speaker Notes:", self.stats['speaker_notes']),
        ]
        
        for row_num, (label, value) in enumerate(metadata, 3):
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            
            if label and label.endswith(":"):
                ws[f'A{row_num}'].font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description="Generate Excel translation record from extraction and translation JSONs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python translation_record_generator.py extracted.json translated.json
  
  # Specify custom output file
  python translation_record_generator.py extracted.json translated.json -o my_record.xlsx
  
  # The Excel file will contain:
  # - Record ID, Slide Number, Element Type, Element Name
  # - Location, Original Text, Translated Text
  # - Character counts, Text Type, Notes
  # - Summary sheet with statistics
        """
    )
    
    parser.add_argument(
        "extraction_json",
        help="Path to extraction JSON file (original content)"
    )
    parser.add_argument(
        "translation_json",
        help="Path to translation JSON file (translated content)"
    )
    parser.add_argument(
        "-o", "--output",
        help="Output Excel file path (default: translation_record_YYYY-MM-DD.xlsx)"
    )
    
    args = parser.parse_args()
    
    # Check if files exist
    if not os.path.exists(args.extraction_json):
        print(f"Error: Extraction JSON not found: {args.extraction_json}")
        return 1
    
    if not os.path.exists(args.translation_json):
        print(f"Error: Translation JSON not found: {args.translation_json}")
        return 1
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = f"translation_record_{timestamp}.xlsx"
    
    # Generate record
    try:
        generator = TranslationRecordGenerator(args.extraction_json, args.translation_json)
        generator.generate_excel(output_path)
        return 0
    except Exception as e:
        print(f"\n❌ Error generating translation record: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())