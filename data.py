"""
Translation Record Generator

Creates an Excel spreadsheet showing original and translated text side-by-side
from the extracted and translated JSON files.

Usage:
    python data.py extracted.json translated.json output.xlsx
    
Or import as module:
    from data import TranslationRecordGenerator
    generator = TranslationRecordGenerator(extracted_json, translated_json)
    generator.generate_excel(output_path)
"""

import json
import argparse
from typing import Dict, List, Tuple, Any
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TranslationRecordGenerator:
    """Generate Excel translation records from extracted and translated JSON files"""
    
    def __init__(self, extracted_json_path: str, translated_json_path: str):
        """
        Initialize the generator with paths to JSON files.
        
        Args:
            extracted_json_path: Path to extracted JSON (original text)
            translated_json_path: Path to translated JSON (translated text)
        """
        self.extracted_json_path = extracted_json_path
        self.translated_json_path = translated_json_path
        
        print(f"Loading extracted data from: {extracted_json_path}")
        with open(extracted_json_path, 'r', encoding='utf-8') as f:
            self.extracted_data = json.load(f)
        
        print(f"Loading translated data from: {translated_json_path}")
        with open(translated_json_path, 'r', encoding='utf-8') as f:
            self.translated_data = json.load(f)
        
        self.translation_records = []
        
    def extract_text_from_runs(self, paragraphs: List[Dict]) -> str:
        """Extract concatenated text from paragraph runs"""
        if not paragraphs:
            return ""
        
        all_text = []
        for para in paragraphs:
            if "runs" in para:
                para_text = "".join(run.get("text", "") for run in para["runs"])
                if para_text.strip():
                    all_text.append(para_text)
        
        return "\n".join(all_text) if all_text else ""
    
    def extract_text_from_table(self, table_data: Dict) -> List[Tuple[str, str, str]]:
        """
        Extract text from table cells.
        
        Returns:
            List of tuples: (cell_location, original_text, translated_text)
        """
        texts = []
        cells = table_data.get("cells", [])
        
        for cell in cells:
            row = cell.get("row", 0)
            col = cell.get("column", 0)
            cell_location = f"Cell ({row}, {col})"
            
            text = self.extract_text_from_runs(cell.get("paragraphs", []))
            if text.strip():
                texts.append((cell_location, text))
        
        return texts
    
    def extract_text_from_chart(self, chart_data: Dict) -> List[Tuple[str, str]]:
        """
        Extract text from chart elements.
        
        Returns:
            List of tuples: (element_name, text)
        """
        texts = []
        
        # Chart title
        if chart_data.get("title"):
            texts.append(("Chart Title", chart_data["title"]))
        
        # Axis titles
        axis_titles = chart_data.get("axis_titles", {})
        for axis_type, title in axis_titles.items():
            if title:
                texts.append((f"{axis_type.capitalize()} Axis", title))
        
        # Series names
        series_names = chart_data.get("series_names", [])
        for idx, name in enumerate(series_names):
            if name:
                texts.append((f"Series {idx + 1}", name))
        
        # Legend entries
        legend_entries = chart_data.get("legend_entries", [])
        for idx, entry in enumerate(legend_entries):
            if entry:
                texts.append((f"Legend {idx + 1}", entry))
        
        # Data labels
        data_values = chart_data.get("data_values", [])
        for series_idx, series in enumerate(data_values):
            data_labels = series.get("data_labels", [])
            for label in data_labels:
                point_idx = label.get("point_index", 0)
                text = label.get("text", "")
                if text:
                    texts.append((f"Series {series_idx + 1} Point {point_idx}", text))
        
        # Categories
        categories = chart_data.get("categories", [])
        for idx, cat in enumerate(categories):
            if isinstance(cat, str) and cat:
                texts.append((f"Category {idx + 1}", cat))
        
        return texts
    
    def extract_text_from_smartart(self, smartart: Dict) -> List[Tuple[str, str]]:
        """
        Extract text from SmartArt elements.
        
        Returns:
            List of tuples: (element_name, text)
        """
        texts = []
        
        # Node texts
        nodes = smartart.get("nodes", [])
        for idx, node in enumerate(nodes):
            text = node.get("text", "")
            if text:
                level = node.get("level", "?")
                texts.append((f"Node {idx + 1} (Level {level})", text))
        
        # Fallback to texts list if nodes are empty
        if not texts:
            text_list = smartart.get("texts", [])
            for idx, text in enumerate(text_list):
                if text:
                    texts.append((f"SmartArt Text {idx + 1}", text))
        
        return texts
    
    def process_element(self, orig_element: Dict, trans_element: Dict, 
                       slide_num: int, element_idx: int):
        """
        Process a single element and extract translation pairs.
        
        Args:
            orig_element: Original element from extracted JSON
            trans_element: Translated element from translated JSON
            slide_num: Slide number
            element_idx: Element index within slide
        """
        element_type = orig_element.get("element_type", "Unknown")
        shape_name = orig_element.get("shape_name", f"Shape {element_idx}")
        shape_id = orig_element.get("shape_id", "?")
        
        # Handle TextBox and AutoShape
        if element_type in ["TextBox", "AutoShape"]:
            orig_text = self.extract_text_from_runs(orig_element.get("paragraphs", []))
            trans_text = self.extract_text_from_runs(trans_element.get("paragraphs", []))
            
            if orig_text.strip() and trans_text.strip():
                self.translation_records.append({
                    "Slide": slide_num,
                    "Element Type": element_type,
                    "Shape Name": shape_name,
                    "Shape ID": shape_id,
                    "Location": "Main Text",
                    "Original Text": orig_text,
                    "Translated Text": trans_text,
                    "Character Count (Original)": len(orig_text),
                    "Character Count (Translated)": len(trans_text)
                })
        
        # Handle Table
        elif element_type == "Table":
            orig_cells = self.extract_text_from_table(orig_element.get("table_data", {}))
            trans_cells = self.extract_text_from_table(trans_element.get("table_data", {}))
            
            # Match cells by index
            for idx, (cell_loc, orig_text) in enumerate(orig_cells):
                if idx < len(trans_cells):
                    _, trans_text = trans_cells[idx]
                    
                    if orig_text.strip() and trans_text.strip():
                        self.translation_records.append({
                            "Slide": slide_num,
                            "Element Type": element_type,
                            "Shape Name": shape_name,
                            "Shape ID": shape_id,
                            "Location": cell_loc,
                            "Original Text": orig_text,
                            "Translated Text": trans_text,
                            "Character Count (Original)": len(orig_text),
                            "Character Count (Translated)": len(trans_text)
                        })
        
        # Handle Chart
        elif element_type == "Chart":
            orig_chart_texts = self.extract_text_from_chart(orig_element.get("chart_data", {}))
            trans_chart_texts = self.extract_text_from_chart(trans_element.get("chart_data", {}))
            
            # Match by index
            for idx, (location, orig_text) in enumerate(orig_chart_texts):
                if idx < len(trans_chart_texts):
                    _, trans_text = trans_chart_texts[idx]
                    
                    if orig_text.strip() and trans_text.strip():
                        self.translation_records.append({
                            "Slide": slide_num,
                            "Element Type": element_type,
                            "Shape Name": shape_name,
                            "Shape ID": shape_id,
                            "Location": location,
                            "Original Text": orig_text,
                            "Translated Text": trans_text,
                            "Character Count (Original)": len(orig_text),
                            "Character Count (Translated)": len(trans_text)
                        })
    
    def process_speaker_notes(self, orig_notes: Dict, trans_notes: Dict, slide_num: int):
        """Process speaker notes"""
        if orig_notes and trans_notes:
            orig_text = orig_notes.get("text", "")
            trans_text = trans_notes.get("text", "")
            
            if orig_text.strip() and trans_text.strip():
                self.translation_records.append({
                    "Slide": slide_num,
                    "Element Type": "Speaker Notes",
                    "Shape Name": "Notes",
                    "Shape ID": "N/A",
                    "Location": "Speaker Notes",
                    "Original Text": orig_text,
                    "Translated Text": trans_text,
                    "Character Count (Original)": len(orig_text),
                    "Character Count (Translated)": len(trans_text)
                })
    
    def process_smartart(self, orig_smartarts: List[Dict], trans_smartarts: List[Dict], 
                        slide_num: int):
        """Process SmartArt elements"""
        for idx, (orig_sa, trans_sa) in enumerate(zip(orig_smartarts, trans_smartarts)):
            orig_texts = self.extract_text_from_smartart(orig_sa)
            trans_texts = self.extract_text_from_smartart(trans_sa)
            
            for text_idx, (location, orig_text) in enumerate(orig_texts):
                if text_idx < len(trans_texts):
                    _, trans_text = trans_texts[text_idx]
                    
                    if orig_text.strip() and trans_text.strip():
                        self.translation_records.append({
                            "Slide": slide_num,
                            "Element Type": "SmartArt",
                            "Shape Name": f"SmartArt {idx + 1}",
                            "Shape ID": "N/A",
                            "Location": location,
                            "Original Text": orig_text,
                            "Translated Text": trans_text,
                            "Character Count (Original)": len(orig_text),
                            "Character Count (Translated)": len(trans_text)
                        })
    
    def generate_records(self):
        """
        Generate all translation records by comparing extracted and translated JSONs.
        """
        print("\nGenerating translation records...")
        
        orig_slides = self.extracted_data.get("slides", [])
        trans_slides = self.translated_data.get("slides", [])
        
        # Process each slide
        for slide_idx, (orig_slide, trans_slide) in enumerate(zip(orig_slides, trans_slides)):
            slide_num = orig_slide.get("slide_number", slide_idx + 1)
            print(f"  Processing slide {slide_num}...")
            
            # Process elements
            orig_elements = orig_slide.get("elements", [])
            trans_elements = trans_slide.get("elements", [])
            
            for elem_idx, (orig_elem, trans_elem) in enumerate(zip(orig_elements, trans_elements)):
                self.process_element(orig_elem, trans_elem, slide_num, elem_idx)
            
            # Process speaker notes
            orig_notes = orig_slide.get("speaker_notes")
            trans_notes = trans_slide.get("speaker_notes")
            if orig_notes and trans_notes:
                self.process_speaker_notes(orig_notes, trans_notes, slide_num)
            
            # Process SmartArt
            orig_smartarts = orig_slide.get("smartart", [])
            trans_smartarts = trans_slide.get("smartart", [])
            if orig_smartarts and trans_smartarts:
                self.process_smartart(orig_smartarts, trans_smartarts, slide_num)
        
        print(f"\n✓ Generated {len(self.translation_records)} translation records")
        return self.translation_records
    
    def format_excel(self, excel_path: str):
        """
        Apply professional formatting to the Excel file.
        
        Args:
            excel_path: Path to the Excel file to format
        """
        print("\nApplying Excel formatting...")
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Alternating row colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Alternating row colors
            fill = light_fill if row_idx % 2 == 0 else PatternFill()
            
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                
                # Wrap text for Original and Translated columns
                if cell.column in [6, 7]:  # Columns F and G (Original Text, Translated Text)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='top')
        
        # Adjust column widths
        column_widths = {
            'A': 8,   # Slide
            'B': 15,  # Element Type
            'C': 20,  # Shape Name
            'D': 10,  # Shape ID
            'E': 20,  # Location
            'F': 50,  # Original Text
            'G': 50,  # Translated Text
            'H': 12,  # Char Count Original
            'I': 12   # Char Count Translated
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save formatted workbook
        wb.save(excel_path)
        print("✓ Formatting applied")
    
    def generate_excel(self, output_path: str):
        """
        Generate the complete Excel translation record.
        
        Args:
            output_path: Path where the Excel file should be saved
        """
        # Generate records
        if not self.translation_records:
            self.generate_records()
        
        if not self.translation_records:
            print("⚠️  No translation records found!")
            return
        
        # Create DataFrame
        df = pd.DataFrame(self.translation_records)
        
        # Sort by slide number and element type
        df = df.sort_values(['Slide', 'Element Type'])
        
        # Save to Excel
        print(f"\nSaving to Excel: {output_path}")
        df.to_excel(output_path, index=False, sheet_name='Translation Record')
        
        # Apply formatting
        self.format_excel(output_path)
        
        # Print statistics
        print("\n" + "=" * 80)
        print("TRANSLATION RECORD GENERATED")
        print("=" * 80)
        print(f"Output file: {output_path}")
        print(f"Total records: {len(self.translation_records)}")
        print(f"Slides processed: {df['Slide'].nunique()}")
        print(f"Element types: {', '.join(df['Element Type'].unique())}")
        
        # Character count statistics
        total_orig_chars = df['Character Count (Original)'].sum()
        total_trans_chars = df['Character Count (Translated)'].sum()
        print(f"Total characters (original): {total_orig_chars:,}")
        print(f"Total characters (translated): {total_trans_chars:,}")
        
        target_lang = self.translated_data.get('target_language', 'Unknown')
        is_rtl = self.translated_data.get('is_rtl', False)
        print(f"Target language: {target_lang}")
        if is_rtl:
            print(f"Text direction: RTL (Right-to-Left)")
        
        print("=" * 80)


def main():
    """Command-line interface for the translation record generator"""
    parser = argparse.ArgumentParser(
        description="Generate Excel translation records from extracted and translated JSON files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate translation record
  python data.py extracted.json translated.json output.xlsx
  
  # Using files in json_bin directory
  python data.py json_bin/file_extracted.json json_bin/file_translated.json translation_record.xlsx
  
  # Auto-detect output filename
  python data.py json_bin/file_extracted.json json_bin/file_translated.json
        """
    )
    
    parser.add_argument(
        "extracted_json",
        help="Path to extracted JSON file (original text)"
    )
    parser.add_argument(
        "translated_json",
        help="Path to translated JSON file (translated text)"
    )
    parser.add_argument(
        "output_excel",
        nargs='?',
        help="Path to output Excel file (optional, auto-generated if not provided)"
    )
    
    args = parser.parse_args()
    
    # Auto-generate output filename if not provided
    if not args.output_excel:
        # Extract base name from translated JSON
        trans_path = Path(args.translated_json)
        base_name = trans_path.stem.replace("_translated", "")
        args.output_excel = f"{base_name}_translation_record.xlsx"
    
    # Check if input files exist
    if not Path(args.extracted_json).exists():
        print(f"❌ Error: Extracted JSON not found: {args.extracted_json}")
        return 1
    
    if not Path(args.translated_json).exists():
        print(f"❌ Error: Translated JSON not found: {args.translated_json}")
        return 1
    
    # Generate translation record
    try:
        generator = TranslationRecordGenerator(args.extracted_json, args.translated_json)
        generator.generate_excel(args.output_excel)
        return 0
    except Exception as e:
        print(f"\n❌ Error generating translation record: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())

